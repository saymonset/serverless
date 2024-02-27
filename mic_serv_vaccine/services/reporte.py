from openpyxl import load_workbook
from datetime import datetime
import json
import tempfile
from flask import Flask, send_file
import os
from services.applyVaccines import  get_applyVaccinesListWithoutJSOn_service


def get_reporte_test_service():
    # Modifica el archivo de Excel y obtiene su ruta temporal
    ruta_archivo_modificado = modificar_archivo_excel()
    # Devuelve el archivo modificado para su descarga
    return send_file(ruta_archivo_modificado, as_attachment=True)

def get_reporte_bydependent_srv(depndentId):
    # Modifica el archivo de Excel y obtiene su ruta temporal
    ruta_archivo_modificado = modificar_archivo_dependentIdexcel(depndentId)
    # Devuelve el archivo modificado para su descarga
    return send_file(ruta_archivo_modificado, as_attachment=True)

def modificar_archivo_dependentIdexcel(depndentId):
      # Nombre del archivo
    
    nombre_archivo = "scheme//children.xlsx"
    # Obtener el path absoluto del archivo
    path_absoluto = os.path.abspath(nombre_archivo)

    # Ruta del archivo original en el paquete de implementación
    ruta_archivo_original = path_absoluto

    # Crea un archivo temporal para trabajar con él
    archivo_temporal = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')

    # Copia el contenido del archivo original al archivo temporal
    with open(ruta_archivo_original, 'rb') as archivo_origen:
        archivo_temporal.write(archivo_origen.read())

    # Cierra el archivo temporal
    archivo_temporal.close()

    # Carga el archivo temporal con openpyxl
    workbook = load_workbook(archivo_temporal.name)
    sheet = workbook['Hoja1']
    # datos = ['11-03-2023', '11-03-2023', '11-05-2023']
    # fila = 3
    # for dato in datos:
    #     celda = sheet.cell(row=fila, column=5)
    #     celda.value = dato
    #     print(f'celda.value = {celda.value}')
    #     fila += 1
      # Convertir la cadena de texto en una lista de objetos
    data =  get_applyVaccinesListWithoutJSOn_service (1000, 0, depndentId);
    dosis_list = json.loads(data)
    # Recorrer la lista de dosis y obtener cada dosis
    for dosis in dosis_list:
        columReporte = dosis['dosis'].get('columReporte', 'N/A')
        rowReporte = dosis['dosis'].get('rowReporte', 'N/A')
        vaccination_date = dosis.get('vaccination_date', 'N/A')
        if columReporte != 'N/A' and rowReporte != 'N/A':
            columReporte = int(columReporte)
            rowReporte = int(rowReporte)
            celda = sheet.cell(row=rowReporte, column=columReporte)
            if vaccination_date != 'N/A':
                formatted_date = datetime.strptime(vaccination_date, '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%d/%m/%Y')
                celda.value = formatted_date
            else:
                celda.value = vaccination_date
            print(f"columReporte: {columReporte}, rowReporte: {rowReporte}, vaccination_date: {vaccination_date}")    
    # Guarda los cambios en el archivo temporal
    workbook.save(archivo_temporal.name)

    # Cierra el archivo de trabajo
    workbook.close()
  
    print(f'Saliendo el path = {archivo_temporal.name}')
    # Devuelve la ruta del archivo temporal modificado
    return archivo_temporal.name
def modificar_archivo_excel():
      # Nombre del archivo
    nombre_archivo = "scheme//children.xlsx"
    # Obtener el path absoluto del archivo
    path_absoluto = os.path.abspath(nombre_archivo)

    # Ruta del archivo original en el paquete de implementación
    ruta_archivo_original = path_absoluto

    # Crea un archivo temporal para trabajar con él
    archivo_temporal = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')

    # Copia el contenido del archivo original al archivo temporal
    with open(ruta_archivo_original, 'rb') as archivo_origen:
        archivo_temporal.write(archivo_origen.read())

    # Cierra el archivo temporal
    archivo_temporal.close()

    # Carga el archivo temporal con openpyxl
    workbook = load_workbook(archivo_temporal.name)
    sheet = workbook['Hoja1']
    datos = ['11-03-2023', '11-03-2023', '11-05-2023']
    fila = 3
    for dato in datos:
        celda = sheet.cell(row=fila, column=5)
        celda.value = dato
        print(f'celda.value = {celda.value}')
        fila += 1
    # Guarda los cambios en el archivo temporal
    workbook.save(archivo_temporal.name)

    # Cierra el archivo de trabajo
    workbook.close()
  
    print(f'Saliendo el path = {archivo_temporal.name}')
    # Devuelve la ruta del archivo temporal modificado
    return archivo_temporal.name

     
 